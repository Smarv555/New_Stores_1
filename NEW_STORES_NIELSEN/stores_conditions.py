import openpyxl as xl
import os
import re


# DIRECTORY SETUP
def excel_directory(country, excel_name):

    if country.upper() == 'KZ':
        # Excel directory for KZ
        excel_path = r'C:\Users\tsst1002\Desktop\New Stores\KZ'
    elif country.upper() == 'BY':
        # Excel directory for BY
        excel_path = r'C:\Users\tsst1002\Desktop\New Stores\BY'

    excel = os.path.join(excel_path, excel_name)
    return excel


# DIRECTORY SETUP
def notepad_directory(txt_path, txt_name):
    notepad = os.path.join(txt_path, txt_name)
    return notepad


# Open excel function
def open_xl(file):
    wb = xl.load_workbook(file)
    return wb._sheets, wb.sheetnames


# Shopetypes function
def shoptype(country):

    shoptypes = {}

    if country == 'KZ':

        excel_path = r'G:/Shared drives/ROC_IV/Belarus_Kazakhstan/Templates/BL_KZ_NEW_STORE_TEMPLATES/KZ'
        excel_name = 'KZ_New_Store_Template_SCANNING.xlsx'
        excel = os.path.join(excel_path, excel_name)

        wb = xl.load_workbook(excel)
        sheet = wb._sheets[7]

        for row in range(2, 50):

            ac_area = sheet.cell(row=row, column=1).value
            ac_shoptype = sheet.cell(row=row, column=2).value

            if ac_area:
                shoptypes[ac_area] = ac_shoptype
            else:
                break

    elif country == 'BY':

        excel_path = r'G:\Shared drives\ROC_IV\Belarus_Kazakhstan\Templates\BL_KZ_NEW_STORE_TEMPLATES\BY'
        excel_name = 'BY_New_Store_Template_SCANNING.xlsx'
        excel = os.path.join(excel_path, excel_name)

        wb = xl.load_workbook(excel)
        sheet = wb._sheets[6]

        for row in range(2, 50):

            ac_area = sheet.cell(row=row, column=1).value
            ac_shoptype = sheet.cell(row=row, column=2).value

            if ac_area:
                shoptypes[ac_area] = ac_shoptype
            else:
                break

    return shoptypes


# Check for cyrillic character in a string
def has_cyrillic(text):
    return bool(re.search(r'[\u0400-\u04FF]', text))


# Check for whitespaces in a string
def has_space(text):
    return bool(re.search(r'\s', text))


# Excel spell check
def spell_check(country, txt_path, txt_name, excel_name):
    spell_ok = True
    excel = excel_directory(country, excel_name)
    notepad = notepad_directory(txt_path, txt_name)
    sheets_list = open_xl(excel)[0]

    with open(notepad, 'w', encoding="utf-8") as file:
        # Sheets cycle
        for sheet in sheets_list:
            # Columns cycle (start, end)
            for column in range(1, 9999):
                # Rows cycle (start, end)
                for row in range(4, 9999):
                    header = sheet.cell(row=2, column=column).value
                    if header == 'AC_SHOPDESCRIPTION' or header == 'DATE' or header == 'AC_RETAILER' or header == '1st Period':
                        continue
                    elif not header:
                        break

                    cell_value = sheet.cell(row=row, column=column).value
                    if has_cyrillic(str(cell_value)) or has_space(str(cell_value)):
                        print(f'SPELLING PROBLEM --> column {header}; row {row}: {cell_value}')
                        file.write(f'column:{header}\trow:{row}\t{cell_value}\n')
                        spell_ok = False

    return spell_ok


# Excel spell check for one sheet only
def spell_check_sheet(country, txt_path, txt_name, excel_name, rows_num, sheet_num):

    spell_ok = True
    excel = excel_directory(country, excel_name)
    notepad = notepad_directory(txt_path, txt_name)
    sheet = open_xl(excel)[0][sheet_num]

    with open(notepad, 'w', encoding="utf-8") as file:

        for column in range(1, 9999):
            for row in range(3, rows_num + 3):
                header = sheet.cell(row=2, column=column).value
                if header == 'AC_SHOPDESCRIPTION' or header == 'DATE' or header == 'AC_RETAILER' or header == '1st Period':
                    continue
                elif not header:
                    break

                cell_value = sheet.cell(row=row, column=column).value
                if has_cyrillic(str(cell_value)) or has_space(str(cell_value)):
                    print(f'SPELLING PROBLEM --> column {header}; row {row}: {cell_value}')
                    file.write(f'column:{header}\trow:{row}\t{cell_value}\n')
                    spell_ok = False

    return spell_ok


# Check for none values in notepad
def none_check(txt_path, txt_name):

    notepad = notepad_directory(txt_path, txt_name)
    with open(notepad, 'r') as file:
        none = False
        for line in file:
            data = line.rstrip()
            if 'None' in data:
                none = True

    if none:
        return ' --- NONE VALUES FOUND!'
    else:
        return ''

if __name__ == '__main__':
    x = spell_check('KZ', 'KZ_NewStores1.xlsx', 0)
    print(x)
    # # User inputs
    # country_input = str(input('Country code: '))
    # excel_name_input = str(input('Excel file name: ')) + '.xlsx'
    # sheet_num_input = int(input('Sheet number (starts from 0): '))
    # rows_num_input = int(input('Rows: '))

