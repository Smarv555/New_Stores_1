import openpyxl as xl
import os

# Notepad file names
insert_txt = 'insert.txt'
delete_txt = 'delete.txt'

# User inputs
country_input = str(input('Country code: '))
excel_name_input = str(input('Excel file name: ')) + '.xlsx'
rows_num_input = int(input('Rows: '))


# DIRECTORY SETUP
def directory_path(country, txt_name, excel_name):
    # Notepad directory
    txt_path = 'C:/Users/tsst1002/Desktop/price_test/'
    notepad = os.path.join(txt_path, txt_name)

    if country.upper() == 'KZ':
        # Excel directory for KZ
        excel_path = 'G:/My Drive/KZ_test/'
    elif country.upper() == 'BY':
        # Excel directory for BY
        excel_path = 'G:/My Drive/BY_test/'

    excel = os.path.join(excel_path, excel_name)
    return notepad, excel


def insert_extraction(country, txt_name, excel_name, rows_num):

    # Directory function - storing the paths in variables
    notepad, excel = directory_path(country, txt_name, excel_name)

    wb = xl.load_workbook(excel)
    sheet = wb._sheets[0]

    cref_list = []

    with open(notepad, 'w') as file:

        # .xlsx rows cycle starting from row 3
        for row in range(2, rows_num + 2):

            # Check for insert comments in Local Decision
            if str(sheet.cell(row=row, column=16).value).upper() == 'INSERT':

                cref = str(sheet.cell(row=row, column=4).value)
                # Check for duplicate crefs
                if cref not in cref_list:
                    cref_list.append(cref)

                    # Columns definition
                    ac_xcodegr = sheet.cell(row=row, column=3).value
                    ac_cref = cref
                    ac_crefsuffix = sheet.cell(row=row, column=17).value
                    ac_crefdescriprion = sheet.cell(row=row, column=18).value if sheet.cell(row=row, column=18).value else ''

                    if sheet.cell(row=row, column=20).value:
                        ac_pricelb = sheet.cell(row=row, column=20).value
                    else:
                        ac_pricelb = sheet.cell(row=row, column=13).value * 0.7

                    if sheet.cell(row=row, column=21).value:
                        nc_priceub = sheet.cell(row=row, column=21).value
                    else:
                        nc_priceub = sheet.cell(row=row, column=13).value * 1.3

                    nc_periodactivefrom = sheet.cell(row=row, column=28).value
                    nc_periodactiveto = sheet.cell(row=row, column=29).value
                    ac_comment = ''
                    nc_conv = sheet.cell(row=row, column=19).value
                    f_nan_key = sheet.cell(row=row, column=11).value

                    # Columns extraction order
                    file.write(
                        f'{ac_xcodegr},'
                        f'{ac_cref},'
                        f'{ac_crefsuffix},'
                        f'{ac_crefdescriprion},'
                        f'{ac_pricelb},'
                        f'{nc_priceub},'
                        f'{nc_periodactivefrom},'
                        f'{nc_periodactiveto},'
                        f'{ac_comment},'
                        f'{nc_conv},'
                        f'{f_nan_key}\n'
                    )

    print(f'Data extracted to {txt_name}')


def delete_extraction(country, txt_name, excel_name, rows_num):

    # Directory function - storing the paths in variables
    notepad, excel = directory_path(country, txt_name, excel_name)

    wb = xl.load_workbook(excel)
    sheet = wb._sheets[0]

    cref_list = []

    with open(notepad, 'w') as file:

        # .xlsx rows cycle starting from row 3
        for row in range(2, rows_num + 2):

            # Check for insert comments in Local Decision
            if str(sheet.cell(row=row, column=16).value).upper() == 'DELETE':

                cref = str(sheet.cell(row=row, column=4).value)
                # Check for duplicate crefs
                if cref not in cref_list:
                    cref_list.append(cref)

                    # Columns definition
                    ac_xcodegr = sheet.cell(row=row, column=3).value
                    ac_cref = cref
                    nc_periodactivefrom = sheet.cell(row=row, column=28).value
                    nc_periodactiveto = sheet.cell(row=row, column=29).value

                    file.write(
                        f'{ac_xcodegr},'
                        f'{ac_cref},'
                        f'{nc_periodactivefrom},'
                        f'{nc_periodactiveto}\n'
                    )

    print(f'Data extracted to {txt_name}')


# Run extraction functions
if __name__ == '__main__':
    # Insert
    insert_extraction(country_input, insert_txt, excel_name_input, rows_num_input)
    # Delete
    delete_extraction(country_input, delete_txt, excel_name_input, rows_num_input)
